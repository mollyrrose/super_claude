"""
Shift years in training data files (CSV + XLSX) by N years.

path-lint-skip: CLI example paths are placeholders

Usage:
    python shift_year.py <folder_path> [--shift N] [--no-backup]

Examples:
    python shift_year.py "C:/path/to/Contoso"           # shift +1 year (default)
    python shift_year.py "C:/path/to/Contoso" --shift 2  # shift +2 years
    python shift_year.py "C:/path/to/Contoso" --shift -1 # shift -1 year
"""

import argparse
import os
import re
import shutil
import sys
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

YEAR_MIN = 2000
YEAR_MAX = 2099


def detect_years_in_csv(filepath: str) -> set[int]:
    """Detect years present in quoted strings of a CSV file by sampling first 100 lines."""
    years = set()
    with open(filepath, "r", encoding="utf-8") as f:
        for i, line in enumerate(f):
            if i >= 100:
                break
            # Find 4-digit numbers inside quoted strings
            for match in re.finditer(r'"[^"]*"', line):
                quoted = match.group()
                for year_match in re.finditer(r'(\d{4})', quoted):
                    year = int(year_match.group())
                    if YEAR_MIN <= year <= YEAR_MAX:
                        years.add(year)
    return years


def shift_csv_content(content: str, old_year: int, new_year: int) -> str:
    """Shift a single year in CSV content, only within quoted strings."""
    old_str = str(old_year)
    new_str = str(new_year)

    def replace_in_quoted(match):
        quoted = match.group()
        return quoted.replace(old_str, new_str)

    return re.sub(r'"[^"]*"', replace_in_quoted, content)


def shift_csv_file(filepath: str, shift: int) -> dict:
    """Shift all detected years in a CSV file. Returns info about what was changed."""
    years = detect_years_in_csv(filepath)
    if not years:
        return {"file": filepath, "status": "skipped", "reason": "no years detected"}

    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()

    # Process years in reverse order (highest first) to prevent double-shifting
    for year in sorted(years, reverse=True):
        new_year = year + shift
        content = shift_csv_content(content, year, new_year)

    with open(filepath, "w", encoding="utf-8", newline="") as f:
        f.write(content)

    new_years = {y + shift for y in years}
    return {
        "file": filepath,
        "status": "shifted",
        "years": f"{sorted(years)} -> {sorted(new_years)}",
    }


def shift_xlsx_file(filepath: str, shift: int) -> dict:
    """Shift year values in an XLSX file."""
    if not HAS_OPENPYXL:
        return {"file": filepath, "status": "skipped", "reason": "openpyxl not installed"}

    wb = openpyxl.load_workbook(filepath)
    years_found = set()
    cells_changed = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, (int, float)) and YEAR_MIN <= cell.value <= YEAR_MAX:
                    old_val = int(cell.value)
                    years_found.add(old_val)

    if not years_found:
        return {"file": filepath, "status": "skipped", "reason": "no years detected"}

    # Process in reverse order to prevent double-shifting
    for year in sorted(years_found, reverse=True):
        new_year = year + shift
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    if isinstance(cell.value, (int, float)) and int(cell.value) == year:
                        cell.value = new_year
                        cells_changed += 1

    wb.save(filepath)

    new_years = {y + shift for y in years_found}
    return {
        "file": filepath,
        "status": "shifted",
        "years": f"{sorted(years_found)} -> {sorted(new_years)}",
        "cells_changed": cells_changed,
    }


def rename_file_with_year(filepath: str, shift: int) -> str | None:
    """Rename file if its name contains a year. Returns new path or None."""
    name = os.path.basename(filepath)
    new_name = name

    # Find years in filename and replace (reverse order)
    years_in_name = sorted(
        {int(m.group()) for m in re.finditer(r'(\d{4})', name)
         if YEAR_MIN <= int(m.group()) <= YEAR_MAX},
        reverse=True,
    )

    for year in years_in_name:
        new_name = new_name.replace(str(year), str(year + shift))

    if new_name != name:
        new_path = os.path.join(os.path.dirname(filepath), new_name)
        os.rename(filepath, new_path)
        return new_path
    return None


def create_backup(folder: str) -> str:
    """Create a backup of the folder contents. Returns backup path."""
    backup_dir = os.path.join(folder, "_backup")
    if os.path.exists(backup_dir):
        shutil.rmtree(backup_dir)
    os.makedirs(backup_dir)

    for f in os.listdir(folder):
        if f == "_backup":
            continue
        src = os.path.join(folder, f)
        if os.path.isfile(src):
            shutil.copy2(src, os.path.join(backup_dir, f))

    return backup_dir


def main():
    parser = argparse.ArgumentParser(description="Shift years in training data files")
    parser.add_argument("folder", help="Folder containing data files")
    parser.add_argument("--shift", type=int, default=1, help="Years to shift (default: +1)")
    parser.add_argument("--no-backup", action="store_true", help="Skip creating backup")
    args = parser.parse_args()

    folder = args.folder
    shift = args.shift

    if not os.path.isdir(folder):
        print(f"Error: '{folder}' is not a valid directory")
        sys.exit(1)

    # Collect files
    csv_files = sorted(Path(folder).glob("*.csv"))
    xlsx_files = sorted(Path(folder).glob("*.xlsx"))

    if not csv_files and not xlsx_files:
        print("No CSV or XLSX files found in the folder.")
        sys.exit(0)

    print(f"Folder: {folder}")
    print(f"Shift: {'+' if shift > 0 else ''}{shift} year(s)")
    print(f"Files found: {len(csv_files)} CSV, {len(xlsx_files)} XLSX")
    print()

    # Backup
    if not args.no_backup:
        backup_path = create_backup(folder)
        print(f"Backup created: {backup_path}")
        print()

    results = []

    # Process CSV files
    for csv_path in csv_files:
        fpath = str(csv_path)
        result = shift_csv_file(fpath, shift)
        # Rename after content shift
        new_path = rename_file_with_year(
            fpath if result["status"] == "skipped" else fpath, shift
        )
        if new_path:
            result["renamed_to"] = os.path.basename(new_path)
        results.append(result)
        print(f"  CSV: {os.path.basename(fpath)}")
        if result["status"] == "shifted":
            print(f"       Years: {result['years']}")
        if new_path:
            print(f"       Renamed -> {os.path.basename(new_path)}")

    # Process XLSX files
    for xlsx_path in xlsx_files:
        fpath = str(xlsx_path)
        result = shift_xlsx_file(fpath, shift)
        results.append(result)
        print(f"  XLSX: {os.path.basename(fpath)}")
        if result["status"] == "shifted":
            print(f"        Years: {result['years']} ({result['cells_changed']} cells)")
        elif result["status"] == "skipped":
            print(f"        Skipped: {result['reason']}")

    print()
    shifted = sum(1 for r in results if r["status"] == "shifted")
    skipped = sum(1 for r in results if r["status"] == "skipped")
    print(f"Done! {shifted} files shifted, {skipped} skipped.")


if __name__ == "__main__":
    main()
